from pylogix import PLC
from datetime import datetime
from openpyxl import load_workbook
import threading
import time

class PeriodicInterval(threading.Thread):
    def __init__(self, task_function, period):
        super().__init__()
        self.daemon = True
        self.task_function = task_function
        self.period = period
        self.i = 0
        self.t0 = time.time()
        self.stopper = 0
        self.start()

    def sleep(self):
        self.i += 1
        delta = self.t0 + self.period * self.i - time.time()
        if delta > 0:
            time.sleep(delta)

    def run(self):
        while self.stopper == 0:
            self.task_function()
            self.sleep()

    def stop(self):
        self.stopper = 1

# Define Defaults
fileName = "LoggerLogix.xlsx"

# Load tag names from Excel file
wb = load_workbook(fileName)
ws = wb.active
# Remove timestamp from tagnames
tag_names = [cell.value for cell in ws[1] if cell.value is not None][:-1]
# Setup PLC comms
plc = PLC()
plc.IPAddress = "192.168.123.100"
plc.ProcessorSlot = 2
plc.SocketTimeout = 1

def log_data():
    # Read tag values
    tag_values = [tag.Value for tag in plc.Read(tag_names)]
    # Append timestamp
    tag_values.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    # Write tag values and timestamp to Excel file
    ws.append(tag_values)

# Create a new periodic thread
tag_thread = PeriodicInterval(log_data, 0.1)

input("Press Enter to exit...")
tag_thread.stop()
# Join the tag thread to wait for it to finish
tag_thread.join(timeout=2)
# CLose PLC Comm link
plc.Close()
# Save Excel File
wb.save(fileName)
